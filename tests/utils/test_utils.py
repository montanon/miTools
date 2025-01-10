import json
import pickle
import re
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from typing import List, Optional
from unittest import TestCase
from unittest.mock import call, mock_open, patch

import numpy as np
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from pandas import DataFrame, Series
from treelib import Tree

from mitools.context.dev_object import Dev, get_dev_var
from mitools.exceptions import ArgumentTypeError, ArgumentValueError
from mitools.utils import (
    AttrDict,
    BitArray,
    LazyDict,
    LazyList,
    add_significance,
    auto_adjust_sheet_columns_width,
    build_dir_tree,
    cached_property,
    can_convert_to,
    check_symmetrical_matrix,
    clean_str,
    decode_string,
    dict_from_kwargs,
    display_env_variables,
    encode_string,
    find_str_line_number_in_text,
    fuzz_ratio,
    fuzz_string_in_string,
    get_numbers_from_str,
    invert_dict,
    iprint,
    iterable_chunks,
    lcs_similarity,
    load_pkl_object,
    pretty_dict_str,
    read_json_file,
    read_text_file,
    remove_chars,
    remove_dataframe_duplicates,
    remove_multiple_spaces,
    replace_prefix,
    sort_dict_keys,
    split_strings,
    store_pkl_object,
    store_signature_in_dev,
    str_is_number,
    stretch_string,
    strip_punctuation,
    unpack_list_of_lists,
    write_json_file,
    write_text_file,
)
from mitools.utils.helper_functions import all_can_be_ints, get_file_encoding
from mitools.utils.helper_objects import _new_attr_dict_


class TestAttrDict(unittest.TestCase):
    def test_init_empty(self):
        ad = AttrDict()
        self.assertEqual(len(ad), 0)
        self.assertEqual(str(ad), "AttrDict{}")

    def test_init_with_dict(self):
        input_dict = {"a": 1, "b": 2}
        ad = AttrDict(input_dict)
        self.assertEqual(len(ad), 2)
        self.assertEqual(ad["a"], 1)
        self.assertEqual(ad["b"], 2)

    def test_init_with_kwargs(self):
        ad = AttrDict(x=10, y=20)
        self.assertEqual(ad.x, 10)
        self.assertEqual(ad.y, 20)

    def test_init_with_sequence_of_tuples(self):
        ad = AttrDict([("key1", "val1"), ("key2", "val2")])
        self.assertEqual(ad.key1, "val1")
        self.assertEqual(ad.key2, "val2")

    def test_attribute_access(self):
        ad = AttrDict({"foo": "bar"})
        self.assertEqual(ad.foo, "bar")
        self.assertEqual(ad["foo"], "bar")
        ad.foo = "baz"
        self.assertEqual(ad.foo, "baz")
        self.assertEqual(ad["foo"], "baz")
        ad["foo"] = "qux"
        self.assertEqual(ad.foo, "qux")
        self.assertEqual(ad["foo"], "qux")

    def test_attribute_error(self):
        ad = AttrDict({"a": 1})
        with self.assertRaises(AttributeError):
            _ = ad.non_existent

    def test_item_access(self):
        ad = AttrDict({"one": 1, "two": 2})
        self.assertEqual(ad["one"], 1)
        self.assertEqual(ad["two"], 2)
        with self.assertRaises(KeyError):
            _ = ad["three"]

    def test_item_assignment(self):
        ad = AttrDict()
        ad["test"] = 123
        self.assertEqual(ad.test, 123)

    def test_item_deletion(self):
        ad = AttrDict({"a": 1, "b": 2})
        del ad["a"]
        self.assertNotIn("a", ad)
        with self.assertRaises(KeyError):
            del ad["a"]

    def test_attribute_deletion(self):
        ad = AttrDict({"x": 10})
        del ad.x
        self.assertNotIn("x", ad)
        with self.assertRaises(KeyError):
            del ad.x

    def test_private_dict_protection(self):
        ad = AttrDict()
        with self.assertRaises(KeyError):
            ad["__private_dict__"] = {}
        with self.assertRaises(AttributeError):
            ad.__private_dict__ = {}

    def test_len(self):
        ad = AttrDict({"a": 1, "b": 2, "c": 3})
        self.assertEqual(len(ad), 3)
        del ad["b"]
        self.assertEqual(len(ad), 2)

    def test_contains(self):
        ad = AttrDict(a=1, b=2)
        self.assertIn("a", ad)
        self.assertNotIn("c", ad)

    def test_keys_values_items(self):
        ad = AttrDict({"a": 1, "b": 2})
        self.assertEqual(set(ad.keys()), {"a", "b"})
        self.assertEqual(set(ad.values()), {1, 2})
        self.assertEqual(set(ad.items()), {("a", 1), ("b", 2)})

    def test_iter(self):
        ad = AttrDict({"x": 100, "y": 200})
        keys = list(iter(ad))
        self.assertIn("x", keys)
        self.assertIn("y", keys)

    def test_update_with_dict(self):
        ad = AttrDict({"a": 1})
        ad.update({"b": 2})
        self.assertEqual(ad.b, 2)
        self.assertEqual(len(ad), 2)

    def test_update_with_kwargs(self):
        ad = AttrDict({"a": 1})
        ad.update(c=3)
        self.assertEqual(ad.c, 3)

    def test_update_with_iterable(self):
        ad = AttrDict({"a": 1})
        ad.update([("b", 2), ("c", 3)])
        self.assertEqual(ad.b, 2)
        self.assertEqual(ad.c, 3)

    def test_clear(self):
        ad = AttrDict(a=1, b=2)
        ad.clear()
        self.assertEqual(len(ad), 0)
        self.assertNotIn("a", ad)

    def test_copy(self):
        ad = AttrDict(a=1, b=2)
        ad_copy = ad.copy()
        self.assertIsNot(ad, ad_copy)
        self.assertEqual(ad_copy.a, 1)
        self.assertEqual(ad_copy.b, 2)
        ad_copy.a = 100
        self.assertEqual(ad.a, 1)
        self.assertEqual(ad_copy.a, 100)

    def test_pop(self):
        ad = AttrDict(a=1, b=2)
        val = ad.pop("a")
        self.assertEqual(val, 1)
        self.assertNotIn("a", ad)

        default_val = ad.pop("missing", "default")
        self.assertEqual(default_val, "default")

    def test_reduce(self):
        ad = AttrDict(a=1, b=2)
        func, args = ad.__reduce__()
        self.assertEqual(func, _new_attr_dict_)
        self.assertEqual(set(args), {("a", 1), ("b", 2)})

    def test_pickle(self):
        ad = AttrDict(x=42, y="foo")
        data = pickle.dumps(ad)
        loaded = pickle.loads(data)
        self.assertEqual(loaded.x, 42)
        self.assertEqual(loaded.y, "foo")
        self.assertIsInstance(loaded, AttrDict)

    def test_repr_str(self):
        ad = AttrDict(a=1, b=2)
        rep = repr(ad)
        self.assertTrue(rep.startswith("AttrDict{"))
        self.assertIn("'a': 1", rep)
        self.assertIn("'b': 2", rep)
        self.assertTrue(rep.endswith("}"))
        self.assertEqual(str(ad), rep)

    def test_dir(self):
        ad = AttrDict(a=1, b=2, non_id_key="val", _private=3)
        directory = dir(ad)
        self.assertIn("a", directory)
        self.assertIn("b", directory)
        self.assertIn("non_id_key", directory)
        self.assertIn("_private", directory)

    def test_setattr(self):
        ad = AttrDict()
        ad.foo = "bar"
        self.assertEqual(ad["foo"], "bar")

    def test_delattr(self):
        ad = AttrDict(foo="bar", baz="qux")
        del ad.foo
        self.assertNotIn("foo", ad)
        with self.assertRaises(KeyError):
            del ad.foo

    def test_invalid_setattr_private_dict(self):
        ad = AttrDict()
        with self.assertRaises(AttributeError):
            ad.__private_dict__ = {}

    def test_reserved_key_error(self):
        ad = AttrDict()
        with self.assertRaises(KeyError):
            ad["__private_dict__"] = 123


class TestStringEncodingDecoding(TestCase):
    def test_decode_string_valid_utf8_bytes(self):
        self.assertEqual(decode_string(b"hello"), "hello")

    def test_decode_string_valid_windows_1252_bytes(self):
        self.assertEqual(decode_string(b"\xe9", encoding="windows-1252"), "é")

    def test_decode_string_utf8_fallback(self):
        self.assertEqual(decode_string(b"\xe9"), "é")  # Decoded as utf-8

    def test_decode_string_invalid_encoding(self):
        result = decode_string(b"\xff\xff", encoding="ascii")
        self.assertEqual(result, "ÿÿ")  # Returns original bytes

    def test_decode_string_str_input(self):
        self.assertEqual(decode_string("already a string"), "already a string")

    def test_non_existent_encoding(self):
        self.assertEqual(decode_string(b"\x81", encoding="windows-1252"), "")

    def test_decode_string_invalid_encoding_type(self):
        with self.assertRaises(TypeError):
            decode_string(b"test", encoding=123)

    def test_encode_string_valid_utf8(self):
        self.assertEqual(encode_string("hello"), b"hello")

    def test_encode_string_valid_windows_1252(self):
        self.assertEqual(encode_string("é", encoding="windows-1252"), b"\xe9")

    def test_encode_string_utf8_fallback(self):
        self.assertEqual(encode_string("€", encoding="ascii"), b"\x80")

    def test_encode_string_invalid_encoding(self):
        result = encode_string("€", encoding="ascii")
        self.assertEqual(
            result, encode_string("€", encoding="windows-1252")
        )  # Returns utf-8 encoded value

    def test_encode_string_bytes_input(self):
        self.assertEqual(encode_string(b"already bytes"), "b'already bytes'")

    def test_encode_string_invalid_encoding_type(self):
        with self.assertRaises(TypeError):
            encode_string("test", encoding=123)

    def test_encode_string_non_utf8_str(self):
        self.assertEqual(encode_string("\x81", encoding="windows-1252"), b"\xc2\x81")

    def test_encode_string_invalid_input_type(self):
        self.assertEqual(encode_string(123), "123")


class TestStripPunctuation(TestCase):
    def test_strip_punctuation_all_true(self):
        self.assertEqual(strip_punctuation("!Hello, World!", all=True), "Hello World")
        self.assertEqual(
            strip_punctuation("...This, is! a test?", all=True), "This is a test"
        )
        self.assertEqual(
            strip_punctuation("No-punctuation-here.", all=True), "Nopunctuationhere"
        )
        self.assertEqual(strip_punctuation("!!!", all=True), "")
        self.assertEqual(strip_punctuation("", all=True), "")
        self.assertEqual(strip_punctuation("1234!@#$", all=True), "1234")

    def test_strip_punctuation_all_false(self):
        self.assertEqual(strip_punctuation("!Hello, World!", all=False), "Hello, World")
        self.assertEqual(
            strip_punctuation("...This, is! a test?", all=False), "This, is! a test"
        )
        self.assertEqual(
            strip_punctuation("No-punctuation-here.", all=False), "No-punctuation-here"
        )
        self.assertEqual(strip_punctuation("!!!", all=False), "")
        self.assertEqual(strip_punctuation("", all=False), "")
        self.assertEqual(strip_punctuation("1234!@#$", all=False), "1234")

    def test_strip_punctuation_whitespace_handling(self):
        self.assertEqual(
            strip_punctuation("   !Hello, World!   ", all=True), "Hello World"
        )
        self.assertEqual(
            strip_punctuation("   ...This, is! a test?   ", all=False),
            "This, is! a test",
        )
        self.assertEqual(
            strip_punctuation("   No-punctuation-here.   ", all=False),
            "No-punctuation-here",
        )

    def test_strip_punctuation_unicode_characters(self):
        self.assertEqual(strip_punctuation("¡Hola, Mundo!", all=True), "¡Hola Mundo")
        self.assertEqual(strip_punctuation("¡Hola, Mundo!", all=False), "¡Hola, Mundo")
        self.assertEqual(
            strip_punctuation("   ¡Hola, Mundo!   ", all=True), "¡Hola Mundo"
        )
        self.assertEqual(
            strip_punctuation("   ¡Hola, Mundo!   ", all=False), "¡Hola, Mundo"
        )

    def test_strip_punctuation_non_ascii(self):
        self.assertEqual(
            strip_punctuation("“Hello” ‘World’!", all=True), "“Hello” ‘World’"
        )
        self.assertEqual(
            strip_punctuation("“Hello” ‘World’!", all=False), "“Hello” ‘World’"
        )
        self.assertEqual(
            strip_punctuation("   “Hello” ‘World’!   ", all=True), "“Hello” ‘World’"
        )
        self.assertEqual(
            strip_punctuation("   “Hello” ‘World’!   ", all=False), "“Hello” ‘World’"
        )

    def test_strip_punctuation_numbers_and_special_cases(self):
        self.assertEqual(strip_punctuation("12345", all=True), "12345")
        self.assertEqual(strip_punctuation("12345", all=False), "12345")
        self.assertEqual(strip_punctuation("12345!!!", all=True), "12345")
        self.assertEqual(strip_punctuation("!!!12345", all=True), "12345")
        self.assertEqual(strip_punctuation("12345!!!", all=False), "12345")
        self.assertEqual(strip_punctuation("!!!12345", all=False), "12345")

    def test_invalid_input_type(self):
        with self.assertRaises(AttributeError):
            strip_punctuation(None)
        with self.assertRaises(AttributeError):
            strip_punctuation(12345)
        with self.assertRaises(AttributeError):
            strip_punctuation(["Hello, World!"], all=True)


class TestLazyList(TestCase):
    def setUp(self):
        class TestableLazyList(LazyList):
            def __init__(self):
                super().__init__()
                self.load_called = False

            def load(self):
                self.load_called = True
                list.extend(self, [1, 2, 3])

        self.TestableLazyList = TestableLazyList

    def test_no_load_before_usage(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        repr_str = repr(d)
        self.assertTrue(d.load_called)
        self.assertIn("1", repr_str)
        self.assertEqual(len(d), 3)

    def test_len_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        length = len(d)
        self.assertTrue(d.load_called)
        self.assertEqual(length, 3)

    def test_iter_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        items = list(d)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(items, [1, 2, 3])

    def test_contains_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        self.assertIn(1, d)  # Should trigger load
        self.assertTrue(d.load_called)

    def test_insert_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        d.insert(0, 0)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d[0], 0)
        self.assertEqual(d[1], 1)

    def test_append_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        d.append(4)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d[-1], 4)
        self.assertEqual(d[0], 1)

    def test_extend_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        d.extend([4, 5])  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d[-2:], [4, 5])
        self.assertEqual(d[:3], [1, 2, 3])

    def test_remove_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        d.remove(2)  # Should trigger load and remove '2'
        self.assertTrue(d.load_called)
        self.assertNotIn(2, d)
        self.assertEqual(d, [1, 3])

    def test_remove_non_existent(self):
        d = self.TestableLazyList()
        _ = len(d)
        with self.assertRaises(ValueError):
            d.remove(99)

    def test_pop_triggers_load(self):
        d = self.TestableLazyList()
        self.assertFalse(d.load_called)
        val = d.pop()  # Should trigger load, default pop removes last
        self.assertTrue(d.load_called)
        self.assertEqual(val, 3)
        self.assertEqual(d, [1, 2])

    def test_pop_with_index(self):
        d = self.TestableLazyList()
        _ = len(d)
        val = d.pop(0)
        self.assertEqual(val, 1)
        self.assertEqual(d, [2, 3])

    def test_pop_non_existent(self):
        d = self.TestableLazyList()
        # Trigger load first
        _ = len(d)
        with self.assertRaises(IndexError):
            d.pop(99)

    def test_multiple_instances(self):
        d1 = self.TestableLazyList()
        d2 = self.TestableLazyList()
        self.assertFalse(d1.load_called)
        self.assertFalse(d2.load_called)
        _ = len(d1)  # trigger load in d1
        self.assertTrue(d1.load_called)
        self.assertFalse(d2.load_called)
        _ = len(d2)  # now trigger load in d2
        self.assertTrue(d2.load_called)

    def test_repr_after_load(self):
        d = self.TestableLazyList()
        _ = len(d)  # trigger load
        rep = repr(d)
        self.assertIn("1", rep)
        self.assertIn("2", rep)
        self.assertIn("3", rep)

    def test_after_load_normal_list_behavior(self):
        d = self.TestableLazyList()
        _ = len(d)  # trigger load
        d.append(10)
        d.insert(0, 0)
        d.remove(3)
        val = d.pop()
        self.assertEqual(val, 10)
        self.assertEqual(d, [0, 1, 2])
        items = list(d)
        self.assertEqual(items, [0, 1, 2])

    def test_subclass_behavior(self):
        class CustomLoadLazyList(LazyList):
            def load(self):
                list.append(self, 100)

        d = CustomLoadLazyList()
        self.assertEqual(len(d), 1)
        self.assertEqual(d[0], 100)

    def test_contains_non_existent_after_load(self):
        d = self.TestableLazyList()
        _ = len(d)  # trigger load
        self.assertNotIn(999, d)

    def test_extend_merge_values(self):
        d = self.TestableLazyList()
        _ = len(d)  # trigger load
        d.extend([4, 5, 6])
        self.assertEqual(d, [1, 2, 3, 4, 5, 6])

    def test_insert_positions(self):
        d = self.TestableLazyList()
        _ = len(d)  # trigger load
        d.insert(0, 0)  # front
        d.insert(2, "x")  # middle
        d.insert(len(d), "end")  # at the end
        self.assertEqual(d, [0, 1, "x", 2, 3, "end"])


class TestLazyDict(TestCase):
    def setUp(self):
        class TestableLazyDict(LazyDict):
            def __init__(self):
                super().__init__()
                self.load_called = False

            def load(self):
                self.load_called = True
                dict_values = {"a": 1, "b": 2, "c": 3}
                for key, value in dict_values.items():
                    dict.__setitem__(self, key, value)

        self.TestableLazyDict = TestableLazyDict

    def test_no_load_before_usage(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        repr_str = repr(d)
        self.assertTrue(d.load_called)
        self.assertIn("a", repr_str)
        self.assertEqual(len(d), 3)

    def test_len_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        length = len(d)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(length, 3)

    def test_iter_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        keys = list(d)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertCountEqual(keys, ["a", "b", "c"])

    def test_contains_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        self.assertIn("a", d)  # Should trigger load
        self.assertTrue(d.load_called)

    def test_getitem_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        value = d["a"]  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(value, 1)

    def test_setitem_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        d["d"] = 4  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d["d"], 4)
        self.assertEqual(d["a"], 1)

    def test_get_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        val = d.get("a")  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(val, 1)

    def test_setdefault_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        d.setdefault("e", 5)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d["e"], 5)
        val = d.setdefault("a", 10)
        self.assertEqual(val, 1)
        self.assertEqual(d["a"], 1)

    def test_items_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        it = d.items()  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertCountEqual(it, [("a", 1), ("b", 2), ("c", 3)])

    def test_keys_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        k = d.keys()  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertCountEqual(k, ["a", "b", "c"])

    def test_values_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        v = d.values()  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertCountEqual(list(v), [1, 2, 3])

    def test_update_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        d.update({"f": 6})  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(d["f"], 6)
        self.assertEqual(d["a"], 1)

    def test_pop_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        val = d.pop("a", None)  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertEqual(val, 1)
        self.assertNotIn("a", d)

    def test_popitem_triggers_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        key, val = d.popitem()  # Should trigger load
        self.assertTrue(d.load_called)
        self.assertIn(key, ["a", "b", "c"])
        self.assertIn(val, [1, 2, 3])
        self.assertNotIn(key, d)

    def test_no_double_load(self):
        d = self.TestableLazyDict()
        self.assertFalse(d.load_called)
        _ = d["a"]  # first trigger
        self.assertTrue(d.load_called)
        # Reset the flag to see if it's called again
        d.load_called = False
        _ = d["b"]  # access another key, should NOT reload
        self.assertFalse(d.load_called)

    def test_multiple_instances(self):
        d1 = self.TestableLazyDict()
        d2 = self.TestableLazyDict()
        self.assertFalse(d1.load_called)
        self.assertFalse(d2.load_called)
        _ = d1["a"]
        self.assertTrue(d1.load_called)
        self.assertFalse(d2.load_called)
        _ = d2["b"]
        self.assertTrue(d2.load_called)

    def test_non_existent_key_after_load(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        with self.assertRaises(KeyError):
            _ = d["not_here"]
        val = d.get("not_here")
        self.assertIsNone(val)

    def test_pop_non_existent_key(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        val = d.pop("not_here", "default")
        self.assertEqual(val, "default")
        with self.assertRaises(KeyError):
            d.pop("still_not_here")

    def test_repr_after_load(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        rep = repr(d)
        self.assertIn("a", rep)
        self.assertIn("b", rep)
        self.assertIn("c", rep)

    def test_update_merge_values(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        d.update({"a": 10, "x": 99})
        self.assertEqual(d["a"], 10)
        self.assertEqual(d["x"], 99)
        self.assertEqual(d["b"], 2)

    def test_setdefault_existing_key(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        original = d.setdefault("a", 999)
        self.assertEqual(original, 1)
        self.assertEqual(d["a"], 1)

    def test_subclass_behavior(self):
        class CustomLoadLazyDict(LazyDict):
            def load(self):
                dict.__setitem__(self, "z", 100)

        d = CustomLoadLazyDict()
        _ = d["z"]
        self.assertEqual(d["z"], 100)

    def test_load_once_then_normal_dict(self):
        d = self.TestableLazyDict()
        _ = d["a"]  # trigger load
        d["g"] = 7
        self.assertEqual(d["g"], 7)
        val = d.pop("b")
        self.assertEqual(val, 2)
        self.assertNotIn("b", d)
        keys = list(d.keys())
        self.assertCountEqual(keys, ["a", "c", "g"])
        items = list(d.items())
        self.assertIn(("a", 1), items)
        self.assertIn(("c", 3), items)
        self.assertIn(("g", 7), items)


class TestCachedProperty(TestCase):
    def setUp(self):
        class Sample:
            def __init__(self, value):
                self._value = value
                self.compute_count = 0

            @cached_property
            def computed_property(self):
                "Sample computed property."
                self.compute_count += 1
                return self._value * 2

        self.Sample = Sample

    def test_cached_property_basic(self):
        obj = self.Sample(10)
        self.assertEqual(obj.computed_property, 20)
        self.assertEqual(obj.computed_property, 20)
        self.assertEqual(obj.compute_count, 1)

    def test_cached_property_caching(self):
        obj = self.Sample(5)
        first_access = obj.computed_property
        second_access = obj.computed_property
        self.assertEqual(first_access, 10)
        self.assertEqual(second_access, 10)
        self.assertEqual(obj.compute_count, 1)

    def test_cached_property_mutability(self):
        obj = self.Sample(3)
        self.assertEqual(obj.computed_property, 6)
        obj.__dict__["computed_property"] = 15
        self.assertEqual(obj.computed_property, 15)
        self.assertEqual(obj.compute_count, 1)

    def test_cached_property_exceptions(self):
        class ExceptionSample:
            def __init__(self, raise_exception):
                self.raise_exception = raise_exception

            @cached_property
            def error_property(self):
                if self.raise_exception:
                    raise ValueError("Intentional error")
                return 42

        obj = ExceptionSample(raise_exception=True)
        with self.assertRaises(ValueError):
            _ = obj.error_property
        obj.raise_exception = False
        self.assertEqual(obj.error_property, 42)

    def test_cached_property_docstring(self):
        obj = self.Sample(4)
        self.assertEqual(
            obj.__class__.computed_property.__doc__, "Sample computed property."
        )

    def test_cached_property_separate_instances(self):
        obj1 = self.Sample(7)
        obj2 = self.Sample(8)
        self.assertEqual(obj1.computed_property, 14)
        self.assertEqual(obj2.computed_property, 16)
        self.assertEqual(obj1.compute_count, 1)
        self.assertEqual(obj2.compute_count, 1)


class TestIterableChunks(TestCase):
    def test_list_input(self):
        iterable = [1, 2, 3, 4, 5, 6]
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [[1, 2], [3, 4], [5, 6]])

    def test_string_input(self):
        iterable = "123456"
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, ["12", "34", "56"])

    def test_tuple_input(self):
        iterable = (1, 2, 3, 4, 5, 6)
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [(1, 2), (3, 4), (5, 6)])

    def test_bytes_input(self):
        iterable = b"123456"
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [b"12", b"34", b"56"])

    def test_invalid_input(self):
        iterable = set([1, 2, 3, 4, 5, 6])
        chunk_size = 2
        with self.assertRaises(TypeError):
            list(iterable_chunks(iterable, chunk_size))


class TestStrIsNumber(TestCase):
    def test_integer_string(self):
        self.assertTrue(str_is_number("123"))

    def test_float_string(self):
        self.assertTrue(str_is_number("123.456"))

    def test_negative_integer_string(self):
        self.assertTrue(str_is_number("-123"))

    def test_negative_float_string(self):
        self.assertTrue(str_is_number("-123.456"))

    def test_non_numeric_string(self):
        self.assertFalse(str_is_number("abc"))

    def test_empty_string(self):
        self.assertFalse(str_is_number(""))


class TestGetNumbersFromStr(TestCase):
    def test_integer_string(self):
        string = "abc 123 def 456"
        self.assertEqual(get_numbers_from_str(string), [123.0, 456.0])

    def test_float_string(self):
        string = "abc 123.456 def 789.012"
        self.assertEqual(get_numbers_from_str(string), [123.456, 789.012])

    def test_negative_number_string(self):
        string = "abc -123 def -456"
        self.assertEqual(get_numbers_from_str(string), [-123.0, -456.0])

    def test_non_numeric_string(self):
        string = "abc def"
        self.assertEqual(get_numbers_from_str(string), [])

    def test_empty_string(self):
        string = ""
        self.assertEqual(get_numbers_from_str(string), [])

    def test_indexed_return(self):
        string = "abc 123 def 456"
        self.assertEqual(get_numbers_from_str(string, 1), 456.0)


class TestRemoveMultipleSpaces(TestCase):
    def test_multiple_spaces(self):
        string = "abc   def   ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_tabs(self):
        string = "abc\t\t\tdef\t\t\tghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_newlines(self):
        string = "abc\n\ndef\n\nghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_mixed_whitespace(self):
        string = "abc \t \n def \t \n ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_no_extra_spaces(self):
        string = "abc def ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")


class TestFindStrLineNumberInText(TestCase):
    def test_substring_at_start(self):
        text = "abc\ndef\nghi"
        substring = "abc"
        self.assertEqual(find_str_line_number_in_text(text, substring), 0)

    def test_substring_in_middle(self):
        text = "abc\ndef\nghi"
        substring = "def"
        self.assertEqual(find_str_line_number_in_text(text, substring), 1)

    def test_substring_at_end(self):
        text = "abc\ndef\nghi"
        substring = "ghi"
        self.assertEqual(find_str_line_number_in_text(text, substring), 2)

    def test_substring_not_found(self):
        text = "abc\ndef\nghi"
        substring = "jkl"
        self.assertEqual(find_str_line_number_in_text(text, substring), None)


class TestReadTextFile(TestCase):
    @patch("builtins.open", new_callable=mock_open, read_data="abc\ndef\nghi")
    def test_read_text_file(self, mock_file):
        text_path = "dummy_path"
        result = read_text_file(text_path)
        self.assertEqual(result, "abc\ndef\nghi")
        mock_file.assert_called_once_with(text_path, "r")


class TestWriteTextFile(TestCase):
    @patch("builtins.open", new_callable=mock_open)
    def test_write_text_file(self, mock_file):
        text_path = "dummy_path"
        text_to_write = "abc\ndef\nghi"
        write_text_file(text_to_write, text_path)
        mock_file.assert_called_once_with(text_path, "w")
        mock_file().write.assert_called_once_with(text_to_write)


class TestReadJsonFile(TestCase):
    def setUp(self):
        self.test_dir = tempfile.TemporaryDirectory()
        self.file_path = Path(self.test_dir.name) / "test.json"

    def tearDown(self):
        self.test_dir.cleanup()

    def test_read_valid_json(self):
        test_data = {"name": "Alice", "age": 30}
        with open(self.file_path, "w") as f:
            json.dump(test_data, f)
        result = read_json_file(self.file_path)
        self.assertEqual(result, test_data)

    def test_read_nonexistent_file(self):
        nonexistent_path = Path(self.test_dir.name) / "nonexistent.json"
        with self.assertRaises(FileNotFoundError):
            read_json_file(nonexistent_path)

    def test_read_invalid_json(self):
        with open(self.file_path, "w") as f:
            f.write("{invalid_json}")
        with self.assertRaises(json.JSONDecodeError):
            read_json_file(self.file_path)


class TestWriteJsonFile(TestCase):
    def setUp(self):
        self.test_dir = tempfile.TemporaryDirectory()
        self.file_path = Path(self.test_dir.name) / "test.json"

    def tearDown(self):
        self.test_dir.cleanup()

    def test_write_valid_json(self):
        test_data = {"name": "Alice", "age": 30}
        write_json_file(test_data, self.file_path)
        with open(self.file_path, "r") as f:
            result = json.load(f)
        self.assertEqual(result, test_data)

    def test_write_empty_json(self):
        test_data = {}
        write_json_file(test_data, self.file_path)
        with open(self.file_path, "r") as f:
            result = json.load(f)
        self.assertEqual(result, test_data)

    def test_write_overwrites_existing_file(self):
        initial_data = {"name": "Bob"}
        with open(self.file_path, "w") as f:
            json.dump(initial_data, f)
        new_data = {"name": "Alice", "age": 30}
        write_json_file(new_data, self.file_path)
        with open(self.file_path, "r") as f:
            result = json.load(f)
        self.assertEqual(result, new_data)


class TestDictFromKwargs(unittest.TestCase):
    def test_no_arguments(self):
        self.assertEqual(dict_from_kwargs(), {})

    def test_one_argument(self):
        self.assertEqual(dict_from_kwargs(a=1), {"a": 1})

    def test_multiple_arguments(self):
        self.assertEqual(dict_from_kwargs(a=1, b=2, c=3), {"a": 1, "b": 2, "c": 3})


class TestLcsSimilarity(unittest.TestCase):
    def test_identical_strings(self):
        self.assertEqual(lcs_similarity("abc", "abc"), 1.0)

    def test_different_strings(self):
        self.assertEqual(lcs_similarity("abc", "def"), 0.0)

    def test_common_subsequence(self):
        self.assertEqual(lcs_similarity("abc", "adc"), 2 / 3)

    def test_empty_string(self):
        self.assertEqual(lcs_similarity("abc", ""), 0.0)
        self.assertEqual(lcs_similarity("", "abc"), 0.0)
        self.assertEqual(lcs_similarity("", ""), 0.0)


class TestFuzzStringInString(unittest.TestCase):
    def test_fuzz_string_in_string_exact_match(self):
        src_string = "Hello World"
        dst_string = "Hello World"
        self.assertTrue(fuzz_string_in_string(src_string, dst_string))

    def test_fuzz_string_in_string_no_match(self):
        src_string = "Hello World"
        dst_string = "Goodbye World"
        self.assertFalse(fuzz_string_in_string(src_string, dst_string))

    def test_fuzz_string_in_string_partial_match_below_threshold(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertFalse(fuzz_string_in_string(src_string, dst_string, 100))

    def test_fuzz_string_in_string_partial_match_above_threshold(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertTrue(fuzz_string_in_string(src_string, dst_string, 50))


class TestFuzzRatio(unittest.TestCase):
    def test_fuzz_ratio_exact_match(self):
        src_string = "Hello World"
        dst_string = "Hello World"
        self.assertEqual(fuzz_ratio(src_string, dst_string), 100)

    def test_fuzz_ratio_no_match(self):
        src_string = "Hello World"
        dst_string = "Goodbye World"
        self.assertEqual(
            fuzz_ratio(src_string, dst_string),
            fuzz.partial_ratio(src_string, dst_string),
        )

    def test_fuzz_ratio_partial_match(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertEqual(
            fuzz_ratio(src_string, dst_string),
            fuzz.partial_ratio(src_string, dst_string),
        )


class TestReplacePrefix(unittest.TestCase):
    def test_replace_prefix(self):
        string = "Hello World"
        prefix = "Hello"
        replacement = "Goodbye"
        self.assertEqual(replace_prefix(string, prefix, replacement), "Goodbye World")

    def test_replace_prefix_no_match(self):
        string = "Hello World"
        prefix = "Goodbye"
        replacement = "Hello"
        self.assertEqual(replace_prefix(string, prefix, replacement), "Hello World")

    def test_replace_prefix_empty_string(self):
        string = ""
        prefix = "Hello"
        replacement = "Goodbye"
        self.assertEqual(replace_prefix(string, prefix, replacement), "")


class TestSplitStrings(unittest.TestCase):
    def test_split_strings(self):
        str_list = ["HelloWorld", "GoodByeWorld"]
        self.assertEqual(
            split_strings(str_list), ["Hello", "World", "Good", "Bye", "World"]
        )

    def test_split_strings_no_capital_letters(self):
        str_list = ["hello", "world"]
        self.assertEqual(split_strings(str_list), ["hello", "world"])

    def test_split_strings_empty_list(self):
        str_list = []
        self.assertEqual(split_strings(str_list), [])


class TestAddSignificance(unittest.TestCase):
    def test_add_significance_very_significant(self):
        row = Series(["Test (0.001)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.001)***")

    def test_add_significance_significant(self):
        row = Series(["Test (0.03)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.03)**")

    def test_add_significance_moderately_significant(self):
        row = Series(["Test (0.07)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.07)*")

    def test_add_significance_not_significant(self):
        row = Series(["Test (0.2)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.2)")


class TestRemoveDataframeDuplicates(unittest.TestCase):
    def setUp(self):
        self.df1 = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        self.df2 = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        self.df3 = DataFrame({"A": [7, 8, 9], "B": [10, 11, 12]})

    def test_remove_dataframe_duplicates(self):
        dfs = [self.df1, self.df2, self.df3]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 2)
        self.assertTrue(unique_dfs[0].equals(self.df1))
        self.assertTrue(unique_dfs[1].equals(self.df3))

    def test_remove_dataframe_duplicates_no_duplicates(self):
        dfs = [self.df1, self.df3]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 2)
        self.assertTrue(unique_dfs[0].equals(self.df1))
        self.assertTrue(unique_dfs[1].equals(self.df3))

    def test_remove_dataframe_duplicates_all_duplicates(self):
        dfs = [self.df1, self.df1, self.df1]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 1)
        self.assertTrue(unique_dfs[0].equals(self.df1))


class TestCanConvertTo(unittest.TestCase):
    def test_can_convert_to_int_from_int(self):
        items = [1, 2, 3]
        self.assertTrue(can_convert_to(items, int))

    def test_can_convert_to_str_from_str(self):
        items = ["1", "2", "3"]
        self.assertTrue(can_convert_to(items, str))

    def test_can_convert_to_float_from_float(self):
        items = [1.0, 2.0, 3.0]
        self.assertTrue(can_convert_to(items, float))

    def test_can_convert_to_bool_from_bool(self):
        items = [True, False, True]
        self.assertTrue(can_convert_to(items, bool))

    def test_can_convert_to_int_from_str(self):
        items = ["1", "2", "3"]
        self.assertTrue(can_convert_to(items, int))

    def test_can_convert_to_str_from_int(self):
        items = [1, 2, 3]
        self.assertTrue(can_convert_to(items, str))

    def test_can_convert_to_float_from_str(self):
        items = ["1.0", "2.0", "3.0"]
        self.assertTrue(can_convert_to(items, float))

    def test_can_convert_to_bool_from_str(self):
        items = ["True", "False", "True"]
        self.assertTrue(can_convert_to(items, bool))

    def test_can_convert_to_int_from_int_fail(self):
        items = [1, 2, 3, "fail"]
        self.assertFalse(can_convert_to(items, int))

    def test_can_convert_to_float_from_float_fail(self):
        items = [1.0, 2.0, 3.0, "fail"]
        self.assertFalse(can_convert_to(items, float))


class TestInvertDict(unittest.TestCase):
    def test_invert_dict(self):
        dictionary = {"a": 1, "b": 2, "c": 3}
        inverted = {1: "a", 2: "b", 3: "c"}
        self.assertEqual(invert_dict(dictionary), inverted)

    def test_invert_dict_empty(self):
        dictionary = {}
        inverted = {}
        self.assertEqual(invert_dict(dictionary), inverted)

    def test_invert_dict_duplicates(self):
        dictionary = {"a": 1, "b": 1, "c": 2}
        inverted = {1: "b", 2: "c"}
        self.assertEqual(invert_dict(dictionary), inverted)


class TestIPrint(unittest.TestCase):
    @patch("builtins.print")
    def test_iprint_string(self, mock_print):
        iprint("Hello World")
        mock_print.assert_called_with("Hello World")

    @patch("builtins.print")
    def test_iprint_list(self, mock_print):
        iprint(["Hello", "World"])
        calls = [call("Hello"), call("World")]
        mock_print.assert_has_calls(calls, any_order=True)

    @patch("builtins.print")
    def test_iprint_splitter(self, mock_print):
        iprint("Hello World", splitter="-")
        calls = [call("-" * 40), call("Hello World"), call("-" * 40)]
        mock_print.assert_has_calls(calls, any_order=True)

    @patch("builtins.print")
    def test_iprint_color(self, mock_print):
        iprint("Hello World", c="red")
        mock_print.assert_called_with("\033[91mHello World\033[0m")


class TestCheckSymmetricalMatrix(unittest.TestCase):
    def test_check_symmetrical_matrix_symmetrical(self):
        a = np.array([[1, 2, 3], [2, 1, 4], [3, 4, 1]])
        self.assertTrue(check_symmetrical_matrix(a))

    def test_check_symmetrical_matrix_not_symmetrical(self):
        a = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
        self.assertFalse(check_symmetrical_matrix(a))

    def test_check_symmetrical_matrix_symmetrical_with_tolerance(self):
        a = np.array([[1, 2, 3], [2, 1, 4.0001], [3, 4, 1]])
        self.assertTrue(check_symmetrical_matrix(a, rtol=1e-04))

    def test_check_symmetrical_matrix_not_symmetrical_with_tolerance(self):
        a = np.array([[1, 2, 3], [2, 1, 4.1], [3, 4, 1]])
        self.assertFalse(check_symmetrical_matrix(a, rtol=1e-04))


class TestRemoveChars(unittest.TestCase):
    def test_basic_removal(self):
        self.assertEqual(remove_chars("Hello, World!", "lo"), "He, Wrd!")

    def test_no_chars_to_remove(self):
        self.assertEqual(remove_chars("Hello, World!", ""), "Hello, World!")

    def test_remove_all_characters(self):
        self.assertEqual(remove_chars("Hello, World!", "Helo, Wrd!"), "")

    def test_string_with_no_matching_characters(self):
        self.assertEqual(remove_chars("Hello, World!", "abc"), "Hello, World!")

    def test_empty_string_input(self):
        self.assertEqual(remove_chars("", "abc"), "")

    def test_special_characters(self):
        self.assertEqual(remove_chars("H@#llo, W$rld!", "@#$"), "Hllo, Wrld!")


class TestStorePklObject(unittest.TestCase):
    def setUp(self):
        self.test_object = {"key": "value"}
        self.filename = Path("test.pkl")

    def tearDown(self):
        if self.filename.exists():
            self.filename.unlink()

    def test_store_pkl_object(self):
        store_pkl_object(self.test_object, self.filename)
        with open(self.filename, "rb") as input_file:
            loaded_object = pickle.load(input_file)
        self.assertEqual(loaded_object, self.test_object)


class TestLoadPklObject(unittest.TestCase):
    def setUp(self):
        self.test_object = {"key": "value"}
        self.filename = Path("test.pkl")
        store_pkl_object(self.test_object, self.filename)

    def tearDown(self):
        if self.filename.exists():
            self.filename.unlink()

    def test_load_pkl_object(self):
        loaded_object = load_pkl_object(self.filename)
        self.assertEqual(loaded_object, self.test_object)


class TestUnpackListOfLists(unittest.TestCase):
    def test_unpack_list_of_lists(self):
        list_of_lists = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        unpacked = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)

    def test_unpack_list_of_lists_empty(self):
        list_of_lists = []
        unpacked = []
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)

    def test_unpack_list_of_lists_single(self):
        list_of_lists = [[1, 2, 3]]
        unpacked = [1, 2, 3]
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)


class TestBitArray(unittest.TestCase):
    def test_zeros_initialization(self):
        size = 16
        bit_array = BitArray.zeros(size)
        for i in range(size):
            self.assertEqual(bit_array.get_index(i), 0)

    def test_set_and_get_bits(self):
        bit_array = BitArray.zeros(16)
        indices_to_set = [1, 3, 15]
        for index in indices_to_set:
            bit_array[index] = 1
            self.assertEqual(bit_array.get_index(index), 1)

    def test_out_of_range_index(self):
        bit_array = BitArray.zeros(10)
        with self.assertRaises(IndexError):
            bit_array.get_index(10)
        with self.assertRaises(IndexError):
            bit_array[10] = 1

    def test_length(self):
        size = 20
        bit_array = BitArray.zeros(size)
        self.assertEqual(len(bit_array), size)

    def test_repr(self):
        bit_array = BitArray.zeros(8)
        expected_repr = "BitArray([0, 0, 0, 0, 0, 0, 0, 0])"
        self.assertEqual(repr(bit_array), expected_repr)


class TestStretchString(unittest.TestCase):
    def test_normal_case(self):
        self.assertEqual(
            stretch_string("This is a sample string for testing purposes", 10),
            "This is a\nsample\nstring for\ntesting\npurposes",
        )

    def test_no_spaces(self):
        self.assertEqual(
            stretch_string("LongStringWithNoSpaces", 5),
            "LongS\ntring\nWithN\noSpac\nes",
        )

    def test_edge_cases(self):
        self.assertEqual(stretch_string("", 10), "")
        self.assertEqual(stretch_string("Short", 10), "Short")
        self.assertEqual(stretch_string("ExactlyTen", 10), "ExactlyTen")

    def test_whitespace_handling(self):
        self.assertEqual(
            stretch_string("  This   string has  weird spacing ", 10),
            "This\nstring has\nweird\nspacing",
        )

    def test_long_word(self):
        self.assertEqual(
            stretch_string("Supercalifragilisticexpialidocious", 10),
            "Supercalif\nragilistic\nexpialidoc\nious",
        )


class TestAutoAdjustColumnsWidth(unittest.TestCase):
    @patch("openpyxl.utils.get_column_letter", return_value="A")
    def test_auto_adjust_columns_width(self, mock_get_column_letter):
        # Create a mock worksheet with some columns
        wb = Workbook()
        ws = wb.active
        ws.append(["Hello", "World"])
        ws.append(["Longer string here", "Another string"])
        # Call the function to test
        auto_adjust_sheet_columns_width(ws)
        # Check that the width of the columns has been adjusted
        self.assertEqual(ws.column_dimensions["A"].width, 15)
        self.assertEqual(ws.column_dimensions["B"].width, 13)


class TestDisplayEnvVariables(unittest.TestCase):
    def setUp(self):
        self.env_vars = [
            ("small_int", 1),
            ("large_list", list(range(10000))),
            ("string", "hello world"),
            ("large_dict", {i: i for i in range(1000)}),
        ]

    def test_no_large_variables(self):
        threshold_mb = sys.getsizeof(self.env_vars[1][1]) / (1024**2) + 1
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertTrue(df.empty)

    def test_large_variables(self):
        threshold_mb = 0
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertFalse(df.empty)
        self.assertTrue(all(df["Size (MB)"] > threshold_mb))

    def test_edge_cases(self):
        # Empty env_vars
        df_empty = display_env_variables([], 0)
        self.assertTrue(df_empty.empty)
        # Extremely high threshold
        df_high_threshold = display_env_variables(self.env_vars, 1000000)
        self.assertTrue(df_high_threshold.empty)

    def test_different_data_types(self):
        threshold_mb = 0
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertIn("large_list", df["Variable"].values)
        self.assertIn("large_dict", df["Variable"].values)


class TestPrettyDictStr(unittest.TestCase):
    def test_pretty_dict_str(self):
        dictionary = {"key2": "value2", "key1": "value1"}
        pretty_str = pretty_dict_str(dictionary)
        expected_str = json.dumps(dictionary, indent=4, sort_keys=True)
        self.assertEqual(pretty_str, expected_str)

    def test_pretty_dict_str_empty(self):
        dictionary = {}
        pretty_str = pretty_dict_str(dictionary)
        expected_str = json.dumps(dictionary, indent=4, sort_keys=True)
        self.assertEqual(pretty_str, expected_str)


class TestBuildDirTree(unittest.TestCase):
    def setUp(self):
        self.test_dir = Path(tempfile.mkdtemp())
        self.sub_dir = self.test_dir / "sub_dir"
        self.sub_dir.mkdir()
        (self.sub_dir / "file.txt").touch()
        self.tree = Tree()

    def tearDown(self):
        shutil.rmtree(self.test_dir)

    def test_build_dir_tree(self):
        tree = build_dir_tree(self.test_dir, self.tree)
        self.assertEqual(len(tree.all_nodes()), len(self.tree.all_nodes()))
        self.assertTrue(any(node.tag == "sub_dir" for node in tree.all_nodes()))
        self.assertTrue(any(node.tag == "file.txt" for node in tree.all_nodes()))


class TestCleanStr(unittest.TestCase):
    def test_clean_str_no_pattern(self):
        string = "Hello, World!"
        result = clean_str(string, None)
        self.assertEqual(result, string)

    def test_clean_str_with_pattern(self):
        string = "Hello, World!"
        pattern = ","
        result = clean_str(string, pattern)
        self.assertEqual(result, "Hello World!")

    def test_clean_str_with_pattern_and_sub_char(self):
        string = "Hello, World!"
        pattern = ","
        sub_char = ";"
        result = clean_str(string, pattern, sub_char)
        self.assertEqual(result, "Hello; World!")


class TestSortDictKeys(TestCase):
    def test_sort_by_keys_ascending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"a": 3, "b": 2, "c": 4, "d": 1}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_sort_by_keys_descending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"d": 1, "c": 4, "b": 2, "a": 3}
        self.assertEqual(sort_dict_keys(input_dict, reverse=True), expected_output)

    def test_sort_by_values_ascending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"d": 1, "b": 2, "a": 3, "c": 4}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: item[1]), expected_output
        )

    def test_sort_by_values_descending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"c": 4, "a": 3, "b": 2, "d": 1}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: item[1], reverse=True),
            expected_output,
        )

    def test_empty_dict(self):
        input_dict = {}
        expected_output = {}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_single_element_dict(self):
        input_dict = {"a": 1}
        expected_output = {"a": 1}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_invalid_input_type(self):
        with self.assertRaises(ArgumentValueError):
            sort_dict_keys(None)

    def test_sort_with_custom_key_function(self):
        input_dict = {"b": "banana", "a": "apple", "c": "cherry"}
        # Custom key function: sort by the length of the values
        expected_output = {"a": "apple", "c": "cherry", "b": "banana"}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: len(item[1])), expected_output
        )


class TestStoreSignatureInDev(TestCase):
    def setUp(self):
        Dev().clear_vars()

    def test_basic_function_args(self):
        @store_signature_in_dev
        def example_func(x: int, y: str):
            return x + len(y)

        result = example_func(42, "test")
        stored_args = get_dev_var("example_func")

        self.assertEqual(stored_args["x"], 42)
        self.assertEqual(stored_args["y"], "test")
        self.assertEqual(result, 46)  # Verify function still works normally

    def test_function_with_defaults(self):
        @store_signature_in_dev
        def func_with_defaults(x: int = 10, y: str = "default"):
            return x + len(y)

        func_with_defaults()
        stored_args = get_dev_var("func_with_defaults")
        self.assertEqual(stored_args["x"], 10)
        self.assertEqual(stored_args["y"], "default")
        func_with_defaults(x=20)
        stored_args = get_dev_var("func_with_defaults")
        self.assertEqual(stored_args["x"], 20)
        self.assertEqual(stored_args["y"], "default")

    def test_function_with_kwargs(self):
        @store_signature_in_dev
        def func_with_kwargs(x: int, **kwargs):
            return x + len(kwargs)

        func_with_kwargs(10, extra="test", another="value")
        stored_args = get_dev_var("func_with_kwargs")
        self.assertEqual(stored_args["x"], 10)
        self.assertEqual(stored_args["kwargs"], {"extra": "test", "another": "value"})

    def test_function_with_args(self):
        @store_signature_in_dev
        def func_with_args(*args):
            return sum(args)

        func_with_args(1, 2, 3)
        stored_args = get_dev_var("func_with_args")
        self.assertEqual(stored_args["args"], (1, 2, 3))

    def test_method_in_class(self):
        class TestClass:
            @store_signature_in_dev
            def test_method(self, x: int, y: Optional[str] = None):
                return x + (len(y) if y else 0)

        obj = TestClass()
        obj.test_method(42, "test")
        stored_args = get_dev_var("test_method")
        self.assertIn("self", stored_args)
        self.assertEqual(stored_args["x"], 42)
        self.assertEqual(stored_args["y"], "test")

    def test_complex_types(self):
        @store_signature_in_dev
        def complex_func(numbers: List[int], text: str = "default"):
            return sum(numbers) + len(text)

        complex_func([1, 2, 3], "test")
        stored_args = get_dev_var("complex_func")
        self.assertEqual(stored_args["numbers"], [1, 2, 3])
        self.assertEqual(stored_args["text"], "test")

    def test_multiple_calls(self):
        @store_signature_in_dev
        def multi_call(x: int):
            return x * 2

        multi_call(10)
        stored_args = get_dev_var("multi_call")
        self.assertEqual(stored_args["x"], 10)
        multi_call(20)
        stored_args = get_dev_var("multi_call")
        self.assertEqual(stored_args["x"], 20)

    def test_nested_functions(self):
        @store_signature_in_dev
        def outer(x: int):
            @store_signature_in_dev
            def inner(y: int):
                return y * 2

            return inner(x + 1)

        result = outer(5)
        outer_args = get_dev_var("outer")
        inner_args = get_dev_var("inner")
        self.assertEqual(outer_args["x"], 5)
        self.assertEqual(inner_args["y"], 6)
        self.assertEqual(result, 12)

    def test_preserves_docstring(self):
        @store_signature_in_dev
        def documented_func(x: int):
            """This is a test docstring."""
            return x

        self.assertEqual(documented_func.__doc__, "This is a test docstring.")

    def test_preserves_function_name(self):
        @store_signature_in_dev
        def named_func(x: int):
            return x

        self.assertEqual(named_func.__name__, "named_func")


if __name__ == "__main__":
    unittest.main()


class TestAllCanBeInts(TestCase):
    def test_all_ints(self):
        self.assertTrue(all_can_be_ints([1, 2, 3]))

    def test_all_strings_representing_ints(self):
        self.assertTrue(all_can_be_ints(["1", "2", "3"]))

    def test_mixed_types(self):
        self.assertTrue(all_can_be_ints(["1", 2, 3.0]))

    def test_non_convertible_item(self):
        self.assertFalse(all_can_be_ints(["1", "a", "3"]))

    def test_with_empty_list(self):
        self.assertTrue(all_can_be_ints([]))

    def test_with_none(self):
        self.assertFalse(all_can_be_ints([None]))

    def test_with_non_numeric_types(self):
        self.assertFalse(all_can_be_ints([1, 2, "three"]))

    def test_with_nested_list(self):
        self.assertFalse(all_can_be_ints([[1, 2], 3]))

    def test_with_boolean_values(self):
        self.assertTrue(all_can_be_ints([True, False]))


class TestGetFileEncoding(TestCase):
    def setUp(self):
        self.temp_file = tempfile.NamedTemporaryFile(delete=False)
        self.temp_file.write("This is a test.".encode("utf-8"))
        self.temp_file.close()

    def tearDown(self):
        Path(self.temp_file.name).unlink()

    def test_detect_utf8_encoding(self):
        encoding = get_file_encoding(self.temp_file.name)
        self.assertEqual(encoding, "utf-8")

    def test_file_not_found(self):
        with self.assertRaises(FileNotFoundError):
            get_file_encoding("non_existent_file.txt")

    @patch("builtins.open", new_callable=mock_open, read_data=b"\x80\x81\x82")
    def test_detect_low_confidence_encoding(self, mock_file):
        with patch(
            "chardet.detect", return_value={"encoding": "iso-8859-1", "confidence": 0.5}
        ):
            encoding = get_file_encoding("dummy_file.txt")
            self.assertEqual(encoding, "utf-8")

    @patch("builtins.open", side_effect=IOError("Permission denied"))
    def test_io_error(self, mock_file):
        with self.assertRaises(IOError):
            get_file_encoding("dummy_file.txt")
