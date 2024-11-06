import json
import os
import pickle
import re
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import TestCase
from unittest.mock import call, mock_open, patch

import numpy as np
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from pandas import DataFrame, Series
from treelib import Tree

from mitools.exceptions import ArgumentTypeError, ArgumentValueError
from mitools.utils import (
    BitArray,
    add_significance,
    auto_adjust_sheet_columns_width,
    build_dir_tree,
    can_convert_to,
    check_symmetrical_matrix,
    clean_str,
    dict_from_kwargs,
    display_env_variables,
    find_str_line_number_in_text,
    fuzz_ratio,
    fuzz_string_in_string,
    get_numbers_from_str,
    invert_dict,
    iprint,
    iterable_chunks,
    lcs_similarity,
    load_pkl_object,
    pretty_dict_str,
    read_text_file,
    remove_chars,
    remove_dataframe_duplicates,
    remove_multiple_spaces,
    replace_prefix,
    sort_dict_keys,
    split_strings,
    store_pkl_object,
    str_is_number,
    stretch_string,
    unpack_list_of_lists,
    validate_args_types,
)


class TestIterableChunks(unittest.TestCase):
    def test_list_input(self):
        iterable = [1, 2, 3, 4, 5, 6]
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [[1, 2], [3, 4], [5, 6]])

    def test_string_input(self):
        iterable = "123456"
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, ["12", "34", "56"])

    def test_tuple_input(self):
        iterable = (1, 2, 3, 4, 5, 6)
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [(1, 2), (3, 4), (5, 6)])

    def test_bytes_input(self):
        iterable = b"123456"
        chunk_size = 2
        result = list(iterable_chunks(iterable, chunk_size))
        self.assertEqual(result, [b"12", b"34", b"56"])

    def test_invalid_input(self):
        iterable = set([1, 2, 3, 4, 5, 6])
        chunk_size = 2
        with self.assertRaises(TypeError):
            list(iterable_chunks(iterable, chunk_size))


class TestStrIsNumber(unittest.TestCase):
    def test_integer_string(self):
        self.assertTrue(str_is_number("123"))

    def test_float_string(self):
        self.assertTrue(str_is_number("123.456"))

    def test_negative_integer_string(self):
        self.assertTrue(str_is_number("-123"))

    def test_negative_float_string(self):
        self.assertTrue(str_is_number("-123.456"))

    def test_non_numeric_string(self):
        self.assertFalse(str_is_number("abc"))

    def test_empty_string(self):
        self.assertFalse(str_is_number(""))


class TestGetNumbersFromStr(unittest.TestCase):
    def test_integer_string(self):
        string = "abc 123 def 456"
        self.assertEqual(get_numbers_from_str(string), [123.0, 456.0])

    def test_float_string(self):
        string = "abc 123.456 def 789.012"
        self.assertEqual(get_numbers_from_str(string), [123.456, 789.012])

    def test_negative_number_string(self):
        string = "abc -123 def -456"
        self.assertEqual(get_numbers_from_str(string), [-123.0, -456.0])

    def test_non_numeric_string(self):
        string = "abc def"
        self.assertEqual(get_numbers_from_str(string), [])

    def test_empty_string(self):
        string = ""
        self.assertEqual(get_numbers_from_str(string), [])

    def test_indexed_return(self):
        string = "abc 123 def 456"
        self.assertEqual(get_numbers_from_str(string, 1), 456.0)


class TestRemoveMultipleSpaces(unittest.TestCase):
    def test_multiple_spaces(self):
        string = "abc   def   ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_tabs(self):
        string = "abc\t\t\tdef\t\t\tghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_newlines(self):
        string = "abc\n\ndef\n\nghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_mixed_whitespace(self):
        string = "abc \t \n def \t \n ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")

    def test_no_extra_spaces(self):
        string = "abc def ghi"
        self.assertEqual(remove_multiple_spaces(string), "abc def ghi")


class TestFindStrLineNumberInText(unittest.TestCase):
    def test_substring_at_start(self):
        text = "abc\ndef\nghi"
        substring = "abc"
        self.assertEqual(find_str_line_number_in_text(text, substring), 0)

    def test_substring_in_middle(self):
        text = "abc\ndef\nghi"
        substring = "def"
        self.assertEqual(find_str_line_number_in_text(text, substring), 1)

    def test_substring_at_end(self):
        text = "abc\ndef\nghi"
        substring = "ghi"
        self.assertEqual(find_str_line_number_in_text(text, substring), 2)

    def test_substring_not_found(self):
        text = "abc\ndef\nghi"
        substring = "jkl"
        self.assertEqual(find_str_line_number_in_text(text, substring), None)


class TestReadTextFile(unittest.TestCase):
    @patch("builtins.open", new_callable=mock_open, read_data="abc\ndef\nghi")
    def test_read_text_file(self, mock_file):
        text_path = "dummy_path"
        result = read_text_file(text_path)
        self.assertEqual(result, "abc\ndef\nghi")
        mock_file.assert_called_once_with(text_path, "r")


class TestDictFromKwargs(unittest.TestCase):
    def test_no_arguments(self):
        self.assertEqual(dict_from_kwargs(), {})

    def test_one_argument(self):
        self.assertEqual(dict_from_kwargs(a=1), {"a": 1})

    def test_multiple_arguments(self):
        self.assertEqual(dict_from_kwargs(a=1, b=2, c=3), {"a": 1, "b": 2, "c": 3})


class TestLcsSimilarity(unittest.TestCase):
    def test_identical_strings(self):
        self.assertEqual(lcs_similarity("abc", "abc"), 1.0)

    def test_different_strings(self):
        self.assertEqual(lcs_similarity("abc", "def"), 0.0)

    def test_common_subsequence(self):
        self.assertEqual(lcs_similarity("abc", "adc"), 2 / 3)

    def test_empty_string(self):
        self.assertEqual(lcs_similarity("abc", ""), 0.0)
        self.assertEqual(lcs_similarity("", "abc"), 0.0)
        self.assertEqual(lcs_similarity("", ""), 0.0)


class TestFuzzStringInString(unittest.TestCase):
    def test_fuzz_string_in_string_exact_match(self):
        src_string = "Hello World"
        dst_string = "Hello World"
        self.assertTrue(fuzz_string_in_string(src_string, dst_string))

    def test_fuzz_string_in_string_no_match(self):
        src_string = "Hello World"
        dst_string = "Goodbye World"
        self.assertFalse(fuzz_string_in_string(src_string, dst_string))

    def test_fuzz_string_in_string_partial_match_below_threshold(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertFalse(fuzz_string_in_string(src_string, dst_string, 100))

    def test_fuzz_string_in_string_partial_match_above_threshold(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertTrue(fuzz_string_in_string(src_string, dst_string, 50))


class TestFuzzRatio(unittest.TestCase):
    def test_fuzz_ratio_exact_match(self):
        src_string = "Hello World"
        dst_string = "Hello World"
        self.assertEqual(fuzz_ratio(src_string, dst_string), 100)

    def test_fuzz_ratio_no_match(self):
        src_string = "Hello World"
        dst_string = "Goodbye World"
        self.assertEqual(
            fuzz_ratio(src_string, dst_string),
            fuzz.partial_ratio(src_string, dst_string),
        )

    def test_fuzz_ratio_partial_match(self):
        src_string = "Hello World"
        dst_string = "Hello"
        self.assertEqual(
            fuzz_ratio(src_string, dst_string),
            fuzz.partial_ratio(src_string, dst_string),
        )


class TestReplacePrefix(unittest.TestCase):
    def test_replace_prefix(self):
        string = "Hello World"
        prefix = "Hello"
        replacement = "Goodbye"
        self.assertEqual(replace_prefix(string, prefix, replacement), "Goodbye World")

    def test_replace_prefix_no_match(self):
        string = "Hello World"
        prefix = "Goodbye"
        replacement = "Hello"
        self.assertEqual(replace_prefix(string, prefix, replacement), "Hello World")

    def test_replace_prefix_empty_string(self):
        string = ""
        prefix = "Hello"
        replacement = "Goodbye"
        self.assertEqual(replace_prefix(string, prefix, replacement), "")


class TestSplitStrings(unittest.TestCase):
    def test_split_strings(self):
        str_list = ["HelloWorld", "GoodByeWorld"]
        self.assertEqual(
            split_strings(str_list), ["Hello", "World", "Good", "Bye", "World"]
        )

    def test_split_strings_no_capital_letters(self):
        str_list = ["hello", "world"]
        self.assertEqual(split_strings(str_list), ["hello", "world"])

    def test_split_strings_empty_list(self):
        str_list = []
        self.assertEqual(split_strings(str_list), [])


class TestAddSignificance(unittest.TestCase):
    def test_add_significance_very_significant(self):
        row = Series(["Test (0.001)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.001)***")

    def test_add_significance_significant(self):
        row = Series(["Test (0.03)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.03)**")

    def test_add_significance_moderately_significant(self):
        row = Series(["Test (0.07)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.07)*")

    def test_add_significance_not_significant(self):
        row = Series(["Test (0.2)"])
        self.assertEqual(row.apply(add_significance)[0], "Test (0.2)")


class TestRemoveDataframeDuplicates(unittest.TestCase):
    def setUp(self):
        self.df1 = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        self.df2 = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        self.df3 = DataFrame({"A": [7, 8, 9], "B": [10, 11, 12]})

    def test_remove_dataframe_duplicates(self):
        dfs = [self.df1, self.df2, self.df3]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 2)
        self.assertTrue(unique_dfs[0].equals(self.df1))
        self.assertTrue(unique_dfs[1].equals(self.df3))

    def test_remove_dataframe_duplicates_no_duplicates(self):
        dfs = [self.df1, self.df3]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 2)
        self.assertTrue(unique_dfs[0].equals(self.df1))
        self.assertTrue(unique_dfs[1].equals(self.df3))

    def test_remove_dataframe_duplicates_all_duplicates(self):
        dfs = [self.df1, self.df1, self.df1]
        unique_dfs = remove_dataframe_duplicates(dfs)
        self.assertEqual(len(unique_dfs), 1)
        self.assertTrue(unique_dfs[0].equals(self.df1))


class TestCanConvertTo(unittest.TestCase):
    def test_can_convert_to_int_from_int(self):
        items = [1, 2, 3]
        self.assertTrue(can_convert_to(items, int))

    def test_can_convert_to_str_from_str(self):
        items = ["1", "2", "3"]
        self.assertTrue(can_convert_to(items, str))

    def test_can_convert_to_float_from_float(self):
        items = [1.0, 2.0, 3.0]
        self.assertTrue(can_convert_to(items, float))

    def test_can_convert_to_bool_from_bool(self):
        items = [True, False, True]
        self.assertTrue(can_convert_to(items, bool))

    def test_can_convert_to_int_from_str(self):
        items = ["1", "2", "3"]
        self.assertTrue(can_convert_to(items, int))

    def test_can_convert_to_str_from_int(self):
        items = [1, 2, 3]
        self.assertTrue(can_convert_to(items, str))

    def test_can_convert_to_float_from_str(self):
        items = ["1.0", "2.0", "3.0"]
        self.assertTrue(can_convert_to(items, float))

    def test_can_convert_to_bool_from_str(self):
        items = ["True", "False", "True"]
        self.assertTrue(can_convert_to(items, bool))

    def test_can_convert_to_int_from_int_fail(self):
        items = [1, 2, 3, "fail"]
        self.assertFalse(can_convert_to(items, int))

    def test_can_convert_to_float_from_float_fail(self):
        items = [1.0, 2.0, 3.0, "fail"]
        self.assertFalse(can_convert_to(items, float))


class TestInvertDict(unittest.TestCase):
    def test_invert_dict(self):
        dictionary = {"a": 1, "b": 2, "c": 3}
        inverted = {1: "a", 2: "b", 3: "c"}
        self.assertEqual(invert_dict(dictionary), inverted)

    def test_invert_dict_empty(self):
        dictionary = {}
        inverted = {}
        self.assertEqual(invert_dict(dictionary), inverted)

    def test_invert_dict_duplicates(self):
        dictionary = {"a": 1, "b": 1, "c": 2}
        inverted = {1: "b", 2: "c"}
        self.assertEqual(invert_dict(dictionary), inverted)


class TestIPrint(unittest.TestCase):
    @patch("builtins.print")
    def test_iprint_string(self, mock_print):
        iprint("Hello World")
        mock_print.assert_called_with("Hello World")

    @patch("builtins.print")
    def test_iprint_list(self, mock_print):
        iprint(["Hello", "World"])
        calls = [call("Hello"), call("World")]
        mock_print.assert_has_calls(calls, any_order=True)

    @patch("builtins.print")
    def test_iprint_splitter(self, mock_print):
        iprint("Hello World", splitter="-")
        calls = [call("-" * 40), call("Hello World"), call("-" * 40)]
        mock_print.assert_has_calls(calls, any_order=True)

    @patch("builtins.print")
    def test_iprint_color(self, mock_print):
        iprint("Hello World", c="red")
        mock_print.assert_called_with("\033[91mHello World\033[0m")


class TestCheckSymmetricalMatrix(unittest.TestCase):
    def test_check_symmetrical_matrix_symmetrical(self):
        a = np.array([[1, 2, 3], [2, 1, 4], [3, 4, 1]])
        self.assertTrue(check_symmetrical_matrix(a))

    def test_check_symmetrical_matrix_not_symmetrical(self):
        a = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
        self.assertFalse(check_symmetrical_matrix(a))

    def test_check_symmetrical_matrix_symmetrical_with_tolerance(self):
        a = np.array([[1, 2, 3], [2, 1, 4.0001], [3, 4, 1]])
        self.assertTrue(check_symmetrical_matrix(a, rtol=1e-04))

    def test_check_symmetrical_matrix_not_symmetrical_with_tolerance(self):
        a = np.array([[1, 2, 3], [2, 1, 4.1], [3, 4, 1]])
        self.assertFalse(check_symmetrical_matrix(a, rtol=1e-04))


class TestRemoveChars(unittest.TestCase):
    def test_basic_removal(self):
        self.assertEqual(remove_chars("Hello, World!", "lo"), "He, Wrd!")

    def test_no_chars_to_remove(self):
        self.assertEqual(remove_chars("Hello, World!", ""), "Hello, World!")

    def test_remove_all_characters(self):
        self.assertEqual(remove_chars("Hello, World!", "Helo, Wrd!"), "")

    def test_string_with_no_matching_characters(self):
        self.assertEqual(remove_chars("Hello, World!", "abc"), "Hello, World!")

    def test_empty_string_input(self):
        self.assertEqual(remove_chars("", "abc"), "")

    def test_special_characters(self):
        self.assertEqual(remove_chars("H@#llo, W$rld!", "@#$"), "Hllo, Wrld!")


class TestStorePklObject(unittest.TestCase):
    def setUp(self):
        self.test_object = {"key": "value"}
        self.filename = "test.pkl"

    def tearDown(self):
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def test_store_pkl_object(self):
        store_pkl_object(self.test_object, self.filename)
        with open(self.filename, "rb") as input_file:
            loaded_object = pickle.load(input_file)
        self.assertEqual(loaded_object, self.test_object)


class TestLoadPklObject(unittest.TestCase):
    def setUp(self):
        self.test_object = {"key": "value"}
        self.filename = "test.pkl"
        store_pkl_object(self.test_object, self.filename)

    def tearDown(self):
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def test_load_pkl_object(self):
        loaded_object = load_pkl_object(self.filename)
        self.assertEqual(loaded_object, self.test_object)


class TestUnpackListOfLists(unittest.TestCase):
    def test_unpack_list_of_lists(self):
        list_of_lists = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        unpacked = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)

    def test_unpack_list_of_lists_empty(self):
        list_of_lists = []
        unpacked = []
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)

    def test_unpack_list_of_lists_single(self):
        list_of_lists = [[1, 2, 3]]
        unpacked = [1, 2, 3]
        self.assertEqual(unpack_list_of_lists(list_of_lists), unpacked)


class TestBitArray(unittest.TestCase):
    def test_zeros_initialization(self):
        size = 16
        bit_array = BitArray.zeros(size)
        for i in range(size):
            self.assertEqual(bit_array.get_index(i), 0)

    def test_set_and_get_bits(self):
        bit_array = BitArray.zeros(16)
        indices_to_set = [1, 3, 15]
        for index in indices_to_set:
            bit_array[index] = 1
            self.assertEqual(bit_array.get_index(index), 1)

    def test_out_of_range_index(self):
        bit_array = BitArray.zeros(10)
        with self.assertRaises(IndexError):
            bit_array.get_index(10)
        with self.assertRaises(IndexError):
            bit_array[10] = 1

    def test_length(self):
        size = 20
        bit_array = BitArray.zeros(size)
        self.assertEqual(len(bit_array), size)

    def test_repr(self):
        bit_array = BitArray.zeros(8)
        expected_repr = "BitArray([0, 0, 0, 0, 0, 0, 0, 0])"
        self.assertEqual(repr(bit_array), expected_repr)


class TestStretchString(unittest.TestCase):
    def test_normal_case(self):
        self.assertEqual(
            stretch_string("This is a sample string for testing purposes", 10),
            "This is a\nsample\nstring for\ntesting\npurposes",
        )

    def test_no_spaces(self):
        self.assertEqual(
            stretch_string("LongStringWithNoSpaces", 5),
            "LongS\ntring\nWithN\noSpac\nes",
        )

    def test_edge_cases(self):
        self.assertEqual(stretch_string("", 10), "")
        self.assertEqual(stretch_string("Short", 10), "Short")
        self.assertEqual(stretch_string("ExactlyTen", 10), "ExactlyTen")

    def test_whitespace_handling(self):
        self.assertEqual(
            stretch_string("  This   string has  weird spacing ", 10),
            "This\nstring has\nweird\nspacing",
        )

    def test_long_word(self):
        self.assertEqual(
            stretch_string("Supercalifragilisticexpialidocious", 10),
            "Supercalif\nragilistic\nexpialidoc\nious",
        )


class TestAutoAdjustColumnsWidth(unittest.TestCase):
    @patch("openpyxl.utils.get_column_letter", return_value="A")
    def test_auto_adjust_columns_width(self, mock_get_column_letter):
        # Create a mock worksheet with some columns
        wb = Workbook()
        ws = wb.active
        ws.append(["Hello", "World"])
        ws.append(["Longer string here", "Another string"])
        # Call the function to test
        auto_adjust_sheet_columns_width(ws)
        # Check that the width of the columns has been adjusted
        self.assertEqual(ws.column_dimensions["A"].width, 15)
        self.assertEqual(ws.column_dimensions["B"].width, 13)


class TestDisplayEnvVariables(unittest.TestCase):
    def setUp(self):
        self.env_vars = [
            ("small_int", 1),
            ("large_list", list(range(10000))),
            ("string", "hello world"),
            ("large_dict", {i: i for i in range(1000)}),
        ]

    def test_no_large_variables(self):
        threshold_mb = sys.getsizeof(self.env_vars[1][1]) / (1024**2) + 1
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertTrue(df.empty)

    def test_large_variables(self):
        threshold_mb = 0
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertFalse(df.empty)
        self.assertTrue(all(df["Size (MB)"] > threshold_mb))

    def test_edge_cases(self):
        # Empty env_vars
        df_empty = display_env_variables([], 0)
        self.assertTrue(df_empty.empty)
        # Extremely high threshold
        df_high_threshold = display_env_variables(self.env_vars, 1000000)
        self.assertTrue(df_high_threshold.empty)

    def test_different_data_types(self):
        threshold_mb = 0
        df = display_env_variables(self.env_vars, threshold_mb)
        self.assertIn("large_list", df["Variable"].values)
        self.assertIn("large_dict", df["Variable"].values)


class TestPrettyDictStr(unittest.TestCase):
    def test_pretty_dict_str(self):
        dictionary = {"key2": "value2", "key1": "value1"}
        pretty_str = pretty_dict_str(dictionary)
        expected_str = json.dumps(dictionary, indent=4, sort_keys=True)
        self.assertEqual(pretty_str, expected_str)

    def test_pretty_dict_str_empty(self):
        dictionary = {}
        pretty_str = pretty_dict_str(dictionary)
        expected_str = json.dumps(dictionary, indent=4, sort_keys=True)
        self.assertEqual(pretty_str, expected_str)


class TestBuildDirTree(unittest.TestCase):
    def setUp(self):
        self.test_dir = Path(tempfile.mkdtemp())
        self.sub_dir = self.test_dir / "sub_dir"
        self.sub_dir.mkdir()
        (self.sub_dir / "file.txt").touch()
        self.tree = Tree()

    def tearDown(self):
        shutil.rmtree(self.test_dir)

    def test_build_dir_tree(self):
        tree = build_dir_tree(self.test_dir, self.tree)
        self.assertEqual(len(tree.all_nodes()), len(self.tree.all_nodes()))
        self.assertTrue(any(node.tag == "sub_dir" for node in tree.all_nodes()))
        self.assertTrue(any(node.tag == "file.txt" for node in tree.all_nodes()))


class TestCleanStr(unittest.TestCase):
    def test_clean_str_no_pattern(self):
        string = "Hello, World!"
        result = clean_str(string, None)
        self.assertEqual(result, string)

    def test_clean_str_with_pattern(self):
        string = "Hello, World!"
        pattern = ","
        result = clean_str(string, pattern)
        self.assertEqual(result, "Hello World!")

    def test_clean_str_with_pattern_and_sub_char(self):
        string = "Hello, World!"
        pattern = ","
        sub_char = ";"
        result = clean_str(string, pattern, sub_char)
        self.assertEqual(result, "Hello; World!")


class TestSortDictKeys(TestCase):
    def test_sort_by_keys_ascending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"a": 3, "b": 2, "c": 4, "d": 1}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_sort_by_keys_descending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"d": 1, "c": 4, "b": 2, "a": 3}
        self.assertEqual(sort_dict_keys(input_dict, reverse=True), expected_output)

    def test_sort_by_values_ascending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"d": 1, "b": 2, "a": 3, "c": 4}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: item[1]), expected_output
        )

    def test_sort_by_values_descending(self):
        input_dict = {"b": 2, "a": 3, "d": 1, "c": 4}
        expected_output = {"c": 4, "a": 3, "b": 2, "d": 1}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: item[1], reverse=True),
            expected_output,
        )

    def test_empty_dict(self):
        input_dict = {}
        expected_output = {}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_single_element_dict(self):
        input_dict = {"a": 1}
        expected_output = {"a": 1}
        self.assertEqual(sort_dict_keys(input_dict), expected_output)

    def test_invalid_input_type(self):
        with self.assertRaises(ArgumentValueError):
            sort_dict_keys(None)

    def test_sort_with_custom_key_function(self):
        input_dict = {"b": "banana", "a": "apple", "c": "cherry"}
        # Custom key function: sort by the length of the values
        expected_output = {"a": "apple", "c": "cherry", "b": "banana"}
        self.assertEqual(
            sort_dict_keys(input_dict, key=lambda item: len(item[1])), expected_output
        )


class TestValidateTypesDecorator(TestCase):
    def test_correct_types_positional_arguments(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y):
            return True

        self.assertTrue(test_func(10, "hello"))

    def test_correct_types_keyword_arguments(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y):
            return True

        self.assertTrue(test_func(x=10, y="hello"))

    def test_incorrect_type_positional_argument(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y):
            return True

        with self.assertRaises(ArgumentTypeError) as context:
            test_func(10, 20)  # y should be a str, not an int
        self.assertIn("Argument 'y' must be of type str", str(context.exception))

    def test_incorrect_type_keyword_argument(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y):
            return True

        with self.assertRaises(ArgumentTypeError) as context:
            test_func(x=10, y=20)  # y should be a str, not an int
        self.assertIn("Argument 'y' must be of type str", str(context.exception))

    def test_missing_argument(self):
        @validate_args_types(x=int, y=str)
        def test_func(x):
            return True

        with self.assertRaises(ArgumentValueError):
            test_func(10)  # Missing argument 'y'

    def test_extra_argument(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y, z):
            return True

        self.assertTrue(test_func(10, "hello", "extra argument"))

    def test_multiple_arguments_different_types(self):
        @validate_args_types(a=int, b=float, c=str)
        def test_func(a, b, c):
            return True

        self.assertTrue(test_func(5, 3.14, "test"))

    def test_multiple_incorrect_arguments(self):
        @validate_args_types(a=int, b=float, c=str)
        def test_func(a, b, c):
            return True

        with self.assertRaises(ArgumentTypeError) as context:
            test_func(5, "not a float", 10)  # b is incorrect
        self.assertIn("Argument 'b' must be of type float", str(context.exception))

    def test_unexpected_argument_name(self):
        @validate_args_types(a=int, b=str)
        def test_func(x, y):
            return True

        with self.assertRaises(ArgumentValueError) as context:
            test_func(5, "hello")
        self.assertIn(
            "Argument 'a' not found in function signature", str(context.exception)
        )

    def test_with_default_values(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y="default"):
            return True

        self.assertTrue(test_func(5))  # y should use the default value, no TypeError

    def test_type_check_on_default_value(self):
        @validate_args_types(x=int, y=str)
        def test_func(x, y="default"):
            return True

        with self.assertRaises(TypeError):
            test_func(
                5, y=10
            )  # y should be a str, not an int, even with default values present

    def test_no_type_validation_when_not_specified(self):
        @validate_args_types(x=int)
        def test_func(x, y):
            return True

        self.assertTrue(
            test_func(5, "anything")
        )  # y has no specified type, so any type is allowed


if __name__ == "__main__":
    unittest.main()
