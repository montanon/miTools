import itertools
import json
import pickle
import re
import string
import sys
from os import PathLike
from pathlib import Path
from typing import (
    Any,
    Callable,
    Dict,
    Generator,
    Iterable,
    List,
    Optional,
    Pattern,
    Tuple,
    Type,
    TypeVar,
    Union,
)

import numpy as np
import openpyxl
import pandas as pd
from fuzzywuzzy import fuzz
from numpy import ndarray
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, Series
from treelib import Tree

from mitools.exceptions import ArgumentValueError

T = TypeVar("T")
COLOR_CODES = {
    "red": "\033[91m",
    "green": "\033[92m",
    "yellow": "\033[93m",
    "blue": "\033[94m",
    "magenta": "\033[95m",
    "cyan": "\033[96m",
    "reset": "\033[0m",  # Reset to default color
}
# Define a cycle of colors
color_cycler = itertools.cycle(COLOR_CODES.keys() - {"reset"})


PUNCTUATION_REGEX = re.compile(
    "[" + "".join(re.escape(p) for p in string.punctuation) + "]"
)


def strip_punctuation(s: str, all: bool = False) -> str:
    if all:
        return PUNCTUATION_REGEX.sub("", s.strip())
    else:
        return s.strip().strip(string.punctuation)


def iterable_chunks(
    iterable: Iterable[T], chunk_size: int
) -> Generator[Iterable[T], None, None]:
    if not isinstance(iterable, (str, list, tuple, bytes)):
        raise TypeError(
            f"Provided iterable of type {type(iterable).__name__} doesn't support slicing."
        )
    for i in range(0, len(iterable), chunk_size):
        yield iterable[i : i + chunk_size]


def str_is_number(string: str) -> bool:
    try:
        float(string)
        return True
    except ValueError:
        return False


def get_numbers_from_str(string: str, n: Optional[int] = None) -> List:
    pattern = r"(-?\d*\.?\d*(?:[eE][-+]?\d+)?)"
    values = [s for s in re.findall(pattern, string.strip()) if s and s != "-"]
    numbers = [float(s) if s != "." else 0 for s in values]
    return numbers[n] if n else numbers


def remove_multiple_spaces(string: str) -> str:
    return re.sub(r"\s+", " ", string)


def find_str_line_number_in_text(text: str, substring: str) -> int:
    lines = text.split("\n")
    for idx, line in enumerate(lines):
        if substring in line:
            return idx


def read_text_file(text_path: PathLike) -> str:
    with open(text_path, "r") as f:
        return f.read()


def write_text_file(text: str, text_path: PathLike) -> None:
    with open(text_path, "w") as f:
        f.write(text)


def read_json_file(json_path: PathLike) -> Dict:
    with open(json_path, "r") as f:
        return json.load(f)


def write_json_file(data: Dict, json_path: PathLike, ensure_ascii: bool = True) -> None:
    with open(json_path, "w") as f:
        json.dump(data, f, indent=4, ensure_ascii=ensure_ascii)


def dict_from_kwargs(**kwargs: Dict[str, Any]) -> Dict:
    return kwargs


def lcs_similarity(s1: str, s2: str) -> float:
    if not s1 or not s2:
        return 0.0
    if s1 in s2 or s2 in s1:
        return min(len(s1), len(s2)) / max(len(s1), len(s2))
    len_s1, len_s2 = len(s1), len(s2)
    if len_s1 < len_s2:
        s1, s2, len_s1, len_s2 = s2, s1, len_s2, len_s1
    curr_row = [0] * (len_s2 + 1)
    for i in range(len_s1):
        prev_val = curr_row[0]
        for j in range(len_s2):
            temp = curr_row[j + 1]
            if s1[i] == s2[j]:
                curr_row[j + 1] = prev_val + 1
            else:
                curr_row[j + 1] = max(curr_row[j], curr_row[j + 1])
            prev_val = temp
    lcs_length = curr_row[-1]
    return lcs_length / max(len_s1, len_s2)


def fuzz_string_in_string(
    src_string: str, dst_string: str, threshold: Optional[int] = 90
) -> bool:
    similarity_score = fuzz_ratio(src_string, dst_string)
    return similarity_score > threshold


def fuzz_ratio(src_string: str, dst_string: str) -> float:
    similarity_score = fuzz.partial_ratio(src_string, dst_string)
    return similarity_score


def replace_prefix(string: str, prefix: Pattern, replacement: str) -> str:
    return re.sub(r"^" + re.escape(prefix), replacement, string)


def split_strings(str_list: List[str]) -> List[str]:
    new_list = []
    for s in str_list:
        new_list += re.split("(?=[A-Z])", s)
    return [s for s in new_list if s]


def add_significance(row: Series) -> Series:
    p_value = float(row.split(" ")[1].replace("(", "").replace(")", ""))
    if p_value < 0.001:
        return row + "***"
    elif p_value < 0.01:
        return row + "**"
    elif p_value < 0.05:
        return row + "*"
    else:
        return row


def remove_dataframe_duplicates(dfs: List[DataFrame]) -> List[DataFrame]:
    unique_dfs = []
    for i in range(len(dfs)):
        if not any(dfs[i].equals(dfs[j]) for j in range(i + 1, len(dfs))):
            unique_dfs.append(dfs[i])
    return unique_dfs


def can_convert_to(items: Iterable, type: Type) -> bool:
    try:
        return all(isinstance(type(item), type) for item in items)
    except ValueError:
        return False


def invert_dict(dictionary: Dict) -> Dict:
    return {value: key for key, value in dictionary.items()}


def iprint(
    iterable: Union[Iterable, str], splitter: Optional[str] = "", c: Optional[str] = ""
):
    if not hasattr(iprint, "color_cycler"):
        iprint.color_cycler = itertools.cycle(COLOR_CODES.keys() - {"reset"})
    color_code = COLOR_CODES.get(
        c, ""
    )  # Get the ANSI escape code for the specified color
    if c == "cycler":
        color_code = COLOR_CODES[next(iprint.color_cycler)]
    else:
        color_code = COLOR_CODES.get(
            c, ""
        )  # Get the ANSI escape code for the specified color
    if isinstance(iterable, str):
        iterable = [iterable]
    elif not isinstance(iterable, Iterable):
        iterable = [repr(iterable)]
    for item in iterable:
        if splitter:
            print(splitter * 40)
        if color_code:
            print(f"{color_code}{item}{COLOR_CODES['reset']}")
        else:
            print(item)
        if splitter:
            print(splitter * 40)


def check_symmetrical_matrix(
    a: ndarray, rtol: Optional[float] = 1e-05, atol: Optional[float] = 1e-08
) -> bool:
    return np.allclose(a, a.T, rtol=rtol, atol=atol)


def remove_chars(input_string: str, chars_to_remove: str) -> str:
    remove_set = set(chars_to_remove)
    return "".join(char for char in input_string if char not in remove_set)


def store_pkl_object(obj, filename):
    with open(filename, "wb") as output_file:
        pickle.dump(obj, output_file)


def load_pkl_object(filename):
    with open(filename, "rb") as input_file:
        obj = pickle.load(input_file)
        return obj


def unpack_list_of_lists(list_of_lists: List[List]) -> List:
    return [item for sub_list in list_of_lists for item in sub_list]


def stretch_string(string: str, length: Optional[int] = 60) -> str:
    string = " ".join(string.split())
    if len(string) > length:
        index = length
        while index >= 0 and string[index] != " ":
            index -= 1
        if index >= 0:
            return string[:index] + "\n" + stretch_string(string[index + 1 :], length)
        else:
            return string[:length] + "\n" + stretch_string(string[length:], length)
    else:
        return string


def auto_adjust_excel_columns_width(excel_path: Path) -> None:
    book = openpyxl.load_workbook(excel_path)
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        auto_adjust_sheet_columns_width(sheet)
    book.save(excel_path)


def auto_adjust_sheet_columns_width(sheet: Worksheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]  # Filter out None values
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1  # Adding a little extra width
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column[0].column)
        ].width = adjusted_width


def display_env_variables(
    env_vars: List[Tuple[str, Any]], threshold_mb: float
) -> DataFrame:
    large_vars = []
    for name, value in env_vars:
        size_mb = sys.getsizeof(value) / (1024**2)
        if size_mb > threshold_mb:
            info = f"Type: {type(value).__name__}, ID: {id(value)}"
            if hasattr(value, "__doc__"):
                doc = str(value.__doc__).split("\n")[0]
                info += f", Doc: {doc[:50]}..."
            large_vars.append((name, size_mb, info))
    df = DataFrame(large_vars, columns=["Variable", "Size (MB)", "Info"])
    df.sort_values(by="Size (MB)", ascending=False, inplace=True)
    return df


def pretty_dict_str(dictionary: Dict) -> str:
    return json.dumps(dictionary, indent=4, sort_keys=True)


def build_dir_tree(
    directory: PathLike, tree: Optional[Tree] = None, parent: Optional[PathLike] = None
) -> Tree:
    if tree is None:
        tree = Tree()
        tree.create_node(directory.name, str(directory))
        parent = str(directory)
    for item in sorted(directory.iterdir()):
        node_id = str(item)
        if item.is_dir():
            tree.create_node(item.name, node_id, parent=parent)
            build_dir_tree(item, tree, parent=node_id)
        else:
            tree.create_node(item.name, node_id, parent=parent)
    return tree


def clean_str(string: str, pattern: Optional[str], sub_char: Optional[str] = "") -> str:
    return re.sub(rf"{pattern}", sub_char, string)


def save_dataframes_to_excel(
    dataframes_dict: Dict[str, DataFrame], filename: PathLike
) -> None:
    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
        for sheet_name, dataframe in dataframes_dict.items():
            dataframe.to_excel(writer, sheet_name=sheet_name)


def read_html_file(file_path: PathLike) -> str:
    with open(file_path, "r", encoding="utf-8") as file:
        return file.read()


def sort_dict_keys(
    input_dict: Dict, key: Callable = None, reverse: bool = False
) -> List:
    try:
        sorted_dict = dict(
            sorted(
                input_dict.items(),
                key=key if key else lambda item: item[0],
                reverse=reverse,
            )
        )
        return sorted_dict
    except Exception as e:
        raise ArgumentValueError(f"An error occured shile sorting the dict: {e}")
